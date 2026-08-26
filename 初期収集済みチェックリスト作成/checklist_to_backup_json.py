#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
checklist_to_backup_json.py
============================
master_to_checklist.py で生成し、なべさんが記入したExcelチェックリストから、
アプリの「一括インポート」機能で読み込める backup JSON (設計書4.5章スキーマ)
を生成する。

【使い方】
  python3 checklist_to_backup_json.py <チェックリストxlsx> <出力json>
"""
import json
import sys
from datetime import datetime, timezone

import openpyxl

VALID_GENRES = {"touge", "kokudo", "gokokuin", "goshouin", "gosenin", "manhole"}


def main():
    if len(sys.argv) != 3:
        print("使い方: python3 checklist_to_backup_json.py <チェックリストxlsx> <出力json>", file=sys.stderr)
        sys.exit(1)

    xlsx_path, output_path = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    today = datetime.now().strftime("%Y-%m-%d")
    genres_out = {}
    total_collected = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == "README":
            continue
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        try:
            idx = {name: header.index(name) for name in
                   ["item_id", "genre", "collected", "collected_date", "memo"]}
        except ValueError:
            print(f"シート'{sheet_name}'は想定フォーマットではないためスキップします。", file=sys.stderr)
            continue

        items = {}
        genre = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["item_id"]] is None:
                continue
            item_id = str(row[idx["item_id"]])
            genre = str(row[idx["genre"]])
            collected_raw = row[idx["collected"]]
            is_collected = collected_raw is not None and str(collected_raw).strip() != ""
            collected_date = row[idx["collected_date"]]
            memo_raw = row[idx["memo"]]
            memo = str(memo_raw).strip() if memo_raw else ""

            if is_collected:
                collected_at = str(collected_date).strip() if collected_date else today
            else:
                collected_at = None

            # 同一item_idが複数行(複数スポット)にまたがる場合(例: 国道の同一路線番号が
            # 複数の販売店に登場する)、いずれかの行でチェックされていれば
            # 収集済みとして扱う(OR結合)。片方の行だけチェックした場合に
            # 別の行の未入力で上書き消去されてしまう事故を防ぐ。
            if item_id in items:
                existing = items[item_id]
                merged_collected = existing["is_collected"] or is_collected
                merged_collected_at = existing["collected_at"] or collected_at
                merged_memo = existing["memo"] or memo
                if existing["memo"] and memo and existing["memo"] != memo:
                    merged_memo = f"{existing['memo']} / {memo}"
                items[item_id] = {
                    "is_collected": merged_collected,
                    "collected_at": merged_collected_at if merged_collected else None,
                    "memo": merged_memo,
                }
            else:
                items[item_id] = {
                    "is_collected": is_collected,
                    "collected_at": collected_at,
                    "memo": memo,
                }

        if genre is None:
            continue
        if genre not in VALID_GENRES:
            print(f"シート'{sheet_name}'のgenre='{genre}'が不正なためスキップします。", file=sys.stderr)
            continue

        genres_out[genre] = {"items": items}
        collected_in_sheet = sum(1 for v in items.values() if v["is_collected"])
        print(f"{genre}: {len(items)}件中 {collected_in_sheet}件を収集済みとして設定")
        total_collected += collected_in_sheet

    backup = {
        "backup_version": "1.0",
        "created_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "device_info": "checklist_to_backup_json.py",
        "genres": genres_out,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(backup, f, ensure_ascii=False, indent=2)

    print(f"\n出力完了: {output_path} (合計 {total_collected}件を収集済みに設定)")


if __name__ == "__main__":
    main()
