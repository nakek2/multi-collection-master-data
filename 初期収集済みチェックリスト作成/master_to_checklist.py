#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
master_to_checklist.py
========================
master_{genre}.json (1つ以上)から、なべさんが「収集済みかどうか」を
Excel上でチェックできる一覧(チェックリスト)を生成する。

生成されたExcelの「collected」列に何か文字を入れる(済/TRUE/1など、
空欄でなければ何でもよい)と「収集済み」として扱われる。
「collected_date」列にYYYY-MM-DD形式で日付を入れると取得日として使われる
(空欄のままなら checklist_to_backup_json.py 実行時点の日付が自動的に使われる)。
「memo」列は任意でメモを入力できる。

item_id / genre 列は変換時の参照用なので変更しないこと。

生成したチェックリストへ入力後、checklist_to_backup_json.py で
アプリの一括インポート用JSONへ変換する。

【重要】一括インポートは「置き換え」であり「追記」ではない。
既にアプリ内でチェック済みのアイテムがある状態でチェックリスト経由の
インポートを行うと、チェックリストに記載の無い(collected未入力の)
アイテムは「未収集」として上書きされてしまう。
このスクリプトは「まだ何もチェックしていない状態からの、まとめての
初期取り込み」を想定している。

【使い方】
  python3 master_to_checklist.py <master_json...> <出力xlsx>

  例(複数ジャンルを1つのExcelにまとめる):
    python3 master_to_checklist.py master_manhole.json master_kokudo.json checklist.xlsx
"""
import json
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill


def build_sheet(wb, genre_master: dict) -> int:
    genre = genre_master["genre"]
    ws = wb.create_sheet(title=genre[:31])  # シート名は31文字制限

    headers = [
        "item_id", "genre", "item_name", "spot_name", "prefecture",
        "district", "address", "collected", "collected_date", "memo",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    rows = []
    for spot in genre_master.get("spots", []):
        for item in spot.get("items", []):
            rows.append([
                item["item_id"],
                genre,
                item["item_name"],
                spot["name"],
                spot["prefecture"],
                spot["district"],
                spot["address"],
                "",  # collected (空欄=未収集、何か入力=収集済み)
                "",  # collected_date
                "",  # memo
            ])

    # 都道府県・地区・アイテム名の順で並べ替え、見つけやすくする。
    rows.sort(key=lambda r: (r[4], r[5], r[2]))
    for r in rows:
        ws.append(r)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 30
    ws.freeze_panes = "A2"
    return len(rows)


def build_readme(wb):
    ws = wb.create_sheet(title="README", index=0)
    lines = [
        "■ 使い方",
        "1. 各ジャンルのシートで、既に持っているアイテムの「collected」列に",
        "   何か文字(例: 済、TRUE、1など)を入力してください。空欄のままなら未収集扱いです。",
        "2. 「collected_date」列に取得日(YYYY-MM-DD)を入力できます。空欄なら",
        "   このチェックリストをJSON変換した日付が自動的に使われます。",
        "3. 「memo」列は任意でメモを入力できます。",
        "4. item_id / genre 列は変更しないでください(自動で参照されます)。",
        "5. 入力が終わったら checklist_to_backup_json.py でこのExcelを",
        "   一括インポート用JSONに変換し、アプリの設定タブ→一括インポートで読み込みます。",
        "",
        "■ 重要な注意",
        "一括インポートは「置き換え」です。既にアプリ内でチェック済みのアイテムが",
        "ある状態でこの手順を使うと、このチェックリストに記載の無いアイテムは",
        "「未収集」として上書きされてしまいます。まだ何もチェックしていない",
        "ジャンルへの、まとめての初期取り込み用として使ってください。",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


def main():
    if len(sys.argv) < 3:
        print("使い方: python3 master_to_checklist.py <master_json...> <出力xlsx>", file=sys.stderr)
        sys.exit(1)

    *master_paths, output_path = sys.argv[1:]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシート削除
    build_readme(wb)

    total = 0
    for path in master_paths:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        count = build_sheet(wb, data)
        total += count
        print(f"{data['genre']}: {count}件")

    wb.save(output_path)
    print(f"\n出力完了: {output_path} (合計 {total}件)")


if __name__ == "__main__":
    main()
